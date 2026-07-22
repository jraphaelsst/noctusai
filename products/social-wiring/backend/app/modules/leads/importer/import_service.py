"""Import orchestration — ``POST /api/leads/import/preview`` (parses,
never writes) and ``POST /api/leads/import/commit`` (parses + upserts +
writes a ``lead_import_batches`` audit row).

Idempotency: re-running commit on the same workbook UPDATES existing
rows (matched by ``(org_id, source_sheet, source_row)``) instead of
duplicating them, per §4's unique index.
"""
from __future__ import annotations

import io
import logging
import uuid
from datetime import date, datetime, timezone
from typing import Any, Optional
from uuid import UUID

import openpyxl

logger = logging.getLogger(__name__)

from app.modules.leads.importer import xlsx_reader
from app.modules.leads.importer.resolvers import resolve_corretor, resolve_origem
from app.modules.leads.services import dimensions_service, leads_service
from app.modules.leads.services.query import backfill_generated_columns, chunked, iter_leads_rows

_SCHEMA = "social_wiring"
_INSERT_CHUNK_SIZE = 500


def _table(client: Any, name: str):
    # `client` is already `social_wiring`-scoped — see
    # `app/modules/leads/deps.py::get_leads_client`'s docstring.
    return client.table(name)


def _load_workbook(file_bytes: bytes):
    return openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)


def _iter_lead_sheets(workbook: Any):
    for name in workbook.sheetnames:
        if name in xlsx_reader.SKIP_SHEETS:
            continue
        yield name, workbook[name]


def _jsonify(value: Any) -> Any:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return value


def _build_payload(
    parsed: "xlsx_reader.ParsedRow",
    sheet_name: str,
    source_map: dict[str, tuple[str, bool]],
    corretor_map: dict[str, str],
) -> dict:
    origem_id, needs_review = resolve_origem(parsed.origem_raw, source_map)
    corretor_id = resolve_corretor(parsed.corretor_raw, corretor_map)
    return {
        "data_entrada": _jsonify(parsed.data_entrada),
        "codigo_raw": parsed.codigo_raw,
        "codigo_imovel": parsed.codigo_imovel,
        "empreendimento": parsed.empreendimento,
        "regiao": parsed.regiao,
        "origem_id": origem_id,
        "origem_raw": parsed.origem_raw,
        "tipo_lead": parsed.tipo_lead,
        "cliente_nome": parsed.cliente_nome,
        "contato": parsed.contato,
        "contato_tipo": parsed.contato_tipo,
        "contato_norm": parsed.contato_norm,
        "corretor_id": corretor_id,
        "corretor_raw": parsed.corretor_raw,
        "anuncio_tier": parsed.anuncio_tier,
        "observacoes": parsed.observacoes,
        "follow_up_data": _jsonify(parsed.follow_up_data),
        "follow_up_nota": parsed.follow_up_nota,
        "needs_review": needs_review,
        "source_sheet": sheet_name,
        "source_row": parsed.source_row,
    }


def _existing_keys_for_org(client: Any, org_id: UUID) -> set[tuple[str, int]]:
    # Paged (`iter_leads_rows`, not a bare `.select()`) — an org with
    # >1000 total leads used to silently show every row past the
    # PostgREST cap as "new" in the preview, undercounting
    # `rows_existing`. See `services/query.py`'s header for the class
    # of bug this belongs to.
    rows = iter_leads_rows(client, org_id, "source_sheet, source_row")
    return {
        (r["source_sheet"], r["source_row"])
        for r in rows
        if r.get("source_sheet") is not None and r.get("source_row") is not None
    }


def _existing_rows_for_sheet(client: Any, org_id: UUID, sheet_name: str) -> dict[int, dict]:
    """``{source_row: {"id": ..., "edited_at": ...}}`` for existing leads
    in this ``(org, sheet)`` — ``commit`` uses ``id`` to decide
    insert-vs-update, and ``edited_at`` to decide whether an update
    would silently overwrite a human's manual correction (see that
    function's docstring + migration 028). Paged (a sheet dense enough
    to cross the 1000-row PostgREST cap used to make ``commit``
    re-INSERT rows it should UPDATE, hitting
    ``uq_sw_leads_org_sheet_row`` and failing the batch after partial
    inserts)."""
    rows = iter_leads_rows(
        client, org_id, "id, source_row, edited_at", source_sheet=sheet_name
    )
    return {r["source_row"]: {"id": r["id"], "edited_at": r.get("edited_at")} for r in rows}


def preview(client: Any, org_id: UUID, file_bytes: bytes, filename: str) -> dict:
    """Parses + resolves against the org's CURRENT alias map; writes
    NOTHING (no ``ensure_default_dimensions`` call here — if the org has
    no dimensions seeded yet, everything correctly shows unmapped)."""
    workbook = _load_workbook(file_bytes)
    source_map, corretor_map = dimensions_service.build_resolution_maps(client, org_id)
    existing_keys = _existing_keys_for_org(client, org_id)

    sheets: list[str] = []
    rows_read = 0
    rows_skipped = 0
    rows_new = 0
    rows_existing = 0
    unmapped_origens: dict[str, int] = {}
    unmapped_corretores: dict[str, int] = {}
    warnings: list[str] = []
    sample_source: list[tuple[str, "xlsx_reader.ParsedRow"]] = []

    for name, worksheet in _iter_lead_sheets(workbook):
        sheets.append(name)
        result = xlsx_reader.parse_sheet(worksheet, name)
        rows_read += result.rows_read
        rows_skipped += result.rows_skipped
        warnings.extend(result.warnings)
        for parsed in result.rows:
            if (name, parsed.source_row) in existing_keys:
                rows_existing += 1
            else:
                rows_new += 1
            origem_id, _ = resolve_origem(parsed.origem_raw, source_map)
            if origem_id is None and parsed.origem_raw:
                unmapped_origens[parsed.origem_raw] = unmapped_origens.get(parsed.origem_raw, 0) + 1
            corretor_id = resolve_corretor(parsed.corretor_raw, corretor_map)
            if corretor_id is None and parsed.corretor_raw:
                unmapped_corretores[parsed.corretor_raw] = (
                    unmapped_corretores.get(parsed.corretor_raw, 0) + 1
                )
            if len(sample_source) < 20:
                sample_source.append((name, parsed))

    refs = leads_service.build_refs(client, org_id)
    now = datetime.now(timezone.utc).isoformat()
    sample = []
    for name, parsed in sample_source:
        payload = _build_payload(parsed, name, source_map, corretor_map)
        row = backfill_generated_columns(
            {
                "id": str(uuid.uuid4()),
                "org_id": str(org_id),
                "created_at": now,
                "updated_at": None,
                **payload,
            }
        )
        sample.append(leads_service.resolve_lead_out(row, refs))

    return {
        "filename": filename,
        "sheets": sheets,
        "rows_read": rows_read,
        "rows_new": rows_new,
        "rows_existing": rows_existing,
        "rows_skipped": rows_skipped,
        "unmapped_origens": [
            {"alias": a, "count": c}
            for a, c in sorted(unmapped_origens.items(), key=lambda kv: -kv[1])
        ],
        "unmapped_corretores": [
            {"alias": a, "count": c}
            for a, c in sorted(unmapped_corretores.items(), key=lambda kv: -kv[1])
        ],
        "sample": sample,
        "warnings": warnings,
    }


def commit(
    client: Any, org_id: UUID, file_bytes: bytes, filename: str, *, created_by: Optional[UUID] = None
) -> dict:
    """Parses + upserts. Ensures the org's canonical dimensions exist
    FIRST (this is the one write-triggering call site for
    ``ensure_default_dimensions`` — see ``migrations/025_leads.sql``'s
    header comment).

    Never overwrites a manually-edited lead
    ─────────────────────────────────────────
    Re-running commit walks EVERY sheet again (a new month's sheet
    being added doesn't change that), not just the ones that changed
    since the last import. Before this fix, an existing row matched by
    ``(org_id, source_sheet, source_row)`` got the FULL spreadsheet
    payload blasted onto it unconditionally — so a human correction made
    via ``PATCH /api/leads/{id}`` (fixed a name, reassigned a corretor,
    cleared ``needs_review``) since the last import was silently
    reverted on the next one, with ``rows_updated`` just going up and no
    diff surfaced. ``leads_service.update_lead`` now stamps
    ``leads.edited_at`` on every real PATCH (migration 028); this
    function checks it via ``_existing_rows_for_sheet`` and, when set,
    skips the update for that row entirely — counted into BOTH
    ``rows_skipped`` (existing counter) and the new
    ``rows_skipped_edited`` (so the UI can show "N preserved because
    manually edited" distinctly from "N skipped, unparseable date"), and
    a reviewable message is appended to ``warnings`` (persisted on the
    batch row — previously ``xlsx_reader.parse_sheet``'s own warnings
    were computed but discarded here; see migration 028's header)."""
    dimensions_service.ensure_default_dimensions(client, org_id)
    source_map, corretor_map = dimensions_service.build_resolution_maps(client, org_id)
    workbook = _load_workbook(file_bytes)

    batch_payload = {
        "id": str(uuid.uuid4()),
        "org_id": str(org_id),
        "filename": filename,
        "sheets": 0,
        "status": "running",
        "started_at": datetime.now(timezone.utc).isoformat(),
        "created_by": str(created_by) if created_by else None,
    }
    resp = _table(client, "lead_import_batches").insert(batch_payload).execute()
    batch_row = list(resp.data or [batch_payload])[0]
    batch_id = batch_row["id"]

    sheets_count = 0
    rows_read = rows_inserted = rows_updated = rows_skipped = rows_flagged = rows_skipped_edited = 0
    warnings: list[str] = []
    try:
        for name, worksheet in _iter_lead_sheets(workbook):
            sheets_count += 1
            result = xlsx_reader.parse_sheet(worksheet, name)
            rows_read += result.rows_read
            rows_skipped += result.rows_skipped
            warnings.extend(result.warnings)

            existing_rows = _existing_rows_for_sheet(client, org_id, name)
            to_insert: list[dict] = []
            to_update: list[tuple[str, dict]] = []
            for parsed in result.rows:
                payload = _build_payload(parsed, name, source_map, corretor_map)
                payload["org_id"] = str(org_id)
                payload["import_batch_id"] = str(batch_id)
                if payload["needs_review"]:
                    rows_flagged += 1
                existing = existing_rows.get(parsed.source_row)
                if existing is None:
                    # Fresh id/created_at — see `leads_service.create_lead`'s
                    # docstring for why this is explicit rather than a DB
                    # default (mock/real-client parity for the id shape).
                    to_insert.append(
                        {
                            "id": str(uuid.uuid4()),
                            "created_at": datetime.now(timezone.utc).isoformat(),
                            "updated_at": None,
                            **payload,
                        }
                    )
                elif existing.get("edited_at"):
                    rows_skipped += 1
                    rows_skipped_edited += 1
                    warnings.append(
                        f"{name}:{parsed.source_row} — preservado (editado "
                        f"manualmente em {existing['edited_at']}), não "
                        f"sobrescrito pela reimportação"
                    )
                else:
                    to_update.append((existing["id"], payload))

            for chunk in chunked(to_insert, _INSERT_CHUNK_SIZE):
                if chunk:
                    _table(client, "leads").insert(chunk).execute()
            rows_inserted += len(to_insert)

            # NOT batched via `.in_(...)` like `dimensions_service`'s
            # relink/reassign fixes — each row here carries its OWN
            # payload (not one uniform value applied to every matching
            # id), so there's no shared update to batch. Bounded anyway:
            # a sheet is a few hundred rows, not the 12k-row scale the
            # relink/reassign paths guard against.
            for existing_id, payload in to_update:
                _table(client, "leads").update(payload).eq("id", existing_id).execute()
            rows_updated += len(to_update)

        finished_at = datetime.now(timezone.utc).isoformat()
        _table(client, "lead_import_batches").update(
            {
                "sheets": sheets_count,
                "rows_read": rows_read,
                "rows_inserted": rows_inserted,
                "rows_updated": rows_updated,
                "rows_skipped": rows_skipped,
                "rows_skipped_edited": rows_skipped_edited,
                "rows_flagged": rows_flagged,
                "warnings": warnings,
                "status": "ok",
                "finished_at": finished_at,
            }
        ).eq("id", batch_id).execute()
    except Exception as exc:  # pragma: no cover — defensive: surface a failed batch, never swallow
        logger.error(
            "lead import commit failed batch_id=%s org_id=%s filename=%s: %s",
            batch_id, org_id, filename, exc, exc_info=True,
        )
        _table(client, "lead_import_batches").update(
            {
                "status": "erro",
                "erro": str(exc),
                "warnings": warnings,
                "finished_at": datetime.now(timezone.utc).isoformat(),
            }
        ).eq("id", batch_id).execute()
        raise

    return get_batch(client, org_id, batch_id)


def get_batch(client: Any, org_id: UUID, batch_id: Any) -> Optional[dict]:
    resp = (
        _table(client, "lead_import_batches")
        .select("*")
        .eq("org_id", str(org_id))
        .eq("id", str(batch_id))
        .execute()
    )
    rows = list(resp.data or [])
    return rows[0] if rows else None


def list_batches(client: Any, org_id: UUID, *, page: int = 1, page_size: int = 20) -> tuple[list[dict], int]:
    resp = (
        _table(client, "lead_import_batches")
        .select("*")
        .eq("org_id", str(org_id))
        .execute()
    )
    rows = list(resp.data or [])
    rows.sort(key=lambda r: r.get("started_at") or "", reverse=True)
    total = len(rows)
    page = max(1, page)
    page_size = max(1, min(page_size, 200))
    start = (page - 1) * page_size
    return rows[start : start + page_size], total


__all__ = ["preview", "commit", "get_batch", "list_batches"]
