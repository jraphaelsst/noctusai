"""Import preview/commit against a SYNTHETIC in-memory workbook (never
real client data — built with openpyxl in this test, never touching
``CONTROLE LEADS (1).xlsx``). Covers: skip-sheet exclusion, multiple
shapes, the retorno-name-prefix leak, an unparseable-date skip, idempotent
re-commit, and needs_review flagging.
"""
from __future__ import annotations

import io
from datetime import date
from uuid import UUID

import openpyxl
import pytest

from noctusai_lib.testing import MockSupabaseClient

from app.modules.leads.importer import import_service
from app.modules.leads.services import dimensions_service

ORG = UUID("00000000-0000-4000-8000-0000000000a1")


def _scoped_mock():
    db = MockSupabaseClient()
    return db.schema("social_wiring")


def _build_workbook() -> bytes:
    wb = openpyxl.Workbook()
    default = wb.active
    wb.remove(default)

    # Skip sheets — must never contribute rows.
    for name in ("RESUMO LEADS", "METRICAS ", "LEADS CORRETORES", "FECHAMENTO ", "RELATORIO DE VENDAS "):
        ws = wb.create_sheet(name)
        ws.append(["not", "real", "lead", "data"])

    # Shape A: NOVO/RETORNO column + tier (JULHO_26-style).
    ws_a = wb.create_sheet("JULHO_26")
    ws_a.append(["Data ", "CODIGO", None, None, "CLIENTE", "CONTATO ", "CORRETORES", "CONTATO", None])
    ws_a.append([date(2026, 7, 1), "ONE9622 Bosque dos Manacás - Km 21", "SENSEYS", "NOVO", "Alice", " 11 99874.5536", "RENATA", None, "obs text"])
    ws_a.append([date(2026, 7, 2), "ONE100", "ZAP", "RETORNO", "Bob", "11988887777", "CINDY", "SIMPLES", None])
    # A row with an unparseable date (skip + count).
    ws_a.append(["not a date at all", "ONE200", "SENSEYS", "NOVO", "Ghost", "11977776666", "RENATA", None, None])

    # Shape B: no NOVO/RETORNO column, no tier, retorno-prefix leak
    # (2024-era sheets).
    ws_b = wb.create_sheet("JULHO_24")
    ws_b.append(["DATA", "CODIGO", "ORIGEM", "CLIENTE", "CONTATO ", "CONSULTOR"])
    ws_b.append(["01.07.24", "ONE300", "SITE", "Retorno Carla", "11 96666.5555", "PRISCILA"])
    ws_b.append(["02.07.24", "CA400", "VR", "Daniela", "11 95555.4444", "WILSON"])
    # Blank origem cell -> needs_review.
    ws_b.append(["03.07.24", "ONE500", None, "Elias", "11 94444.3333", "FERNANDA"])
    # Unmapped/unknown origem spelling -> outro + needs_review.
    ws_b.append(["04.07.24", "ONE600", "SOME BRAND NEW SPELLING NEVER SEEN", "Fabio", "11 93333.2222", "MARCO"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def workbook_bytes():
    return _build_workbook()


class TestPreview:
    def test_excludes_skip_sheets(self, workbook_bytes):
        client = _scoped_mock()
        result = import_service.preview(client, ORG, workbook_bytes, "test.xlsx")
        assert "RESUMO LEADS" not in result["sheets"]
        assert "METRICAS " not in result["sheets"]
        assert set(result["sheets"]) == {"JULHO_26", "JULHO_24"}

    def test_rows_read_and_skipped_counts(self, workbook_bytes):
        client = _scoped_mock()
        result = import_service.preview(client, ORG, workbook_bytes, "test.xlsx")
        # JULHO_26: 3 candidate rows (1 unparseable) + JULHO_24: 4 rows.
        assert result["rows_read"] == 7
        assert result["rows_skipped"] == 1
        assert result["rows_new"] == 6

    def test_writes_nothing(self, workbook_bytes):
        client = _scoped_mock()
        import_service.preview(client, ORG, workbook_bytes, "test.xlsx")
        rows, total = import_service.list_batches(client, ORG)
        assert total == 0
        assert dimensions_service.list_sources(client, ORG) == []

    def test_unmapped_origens_reported(self, workbook_bytes):
        client = _scoped_mock()
        result = import_service.preview(client, ORG, workbook_bytes, "test.xlsx")
        unmapped = {u["alias"] for u in result["unmapped_origens"]}
        # Nothing seeded yet (preview never seeds) -> every raw origem is
        # unmapped, including ones the canonical alias table WOULD
        # resolve after a commit.
        assert "SENSEYS" in unmapped
        assert "SOME BRAND NEW SPELLING NEVER SEEN" in unmapped

    def test_sample_capped_at_20_and_retorno_prefix_stripped(self, workbook_bytes):
        client = _scoped_mock()
        result = import_service.preview(client, ORG, workbook_bytes, "test.xlsx")
        assert len(result["sample"]) <= 20
        carla = next(r for r in result["sample"] if "Carla" in (r["cliente_nome"] or ""))
        assert carla["cliente_nome"] == "Carla"
        assert carla["tipo_lead"] == "retorno"


class TestCommit:
    def test_ensures_default_dimensions_first(self, workbook_bytes):
        client = _scoped_mock()
        assert dimensions_service.list_sources(client, ORG) == []
        import_service.commit(client, ORG, workbook_bytes, "test.xlsx")
        assert len(dimensions_service.list_sources(client, ORG)) > 0

    def test_writes_leads_and_batch_record(self, workbook_bytes):
        client = _scoped_mock()
        batch = import_service.commit(client, ORG, workbook_bytes, "test.xlsx")
        assert batch["status"] == "ok"
        assert batch["sheets"] == 2
        assert batch["rows_read"] == 7
        assert batch["rows_inserted"] == 6
        assert batch["rows_skipped"] == 1

    def test_needs_review_flagged_for_blank_and_unmapped_origem(self, workbook_bytes):
        client = _scoped_mock()
        batch = import_service.commit(client, ORG, workbook_bytes, "test.xlsx")
        # Elias (blank origem) + Fabio (unmapped spelling) -> 2 flagged.
        assert batch["rows_flagged"] == 2

    def test_resolves_known_alias(self, workbook_bytes):
        client = _scoped_mock()
        import_service.commit(client, ORG, workbook_bytes, "test.xlsx")
        from app.modules.leads.services.query import LeadFilters, fetch_filtered

        rows = fetch_filtered(client, ORG, LeadFilters())
        alice = next(r for r in rows if r["cliente_nome"] == "Alice")
        assert alice["origem_raw"] == "SENSEYS"
        assert alice["origem_id"] is not None
        assert alice["needs_review"] is False

    def test_idempotent_recommit_updates_not_duplicates(self, workbook_bytes):
        client = _scoped_mock()
        import_service.commit(client, ORG, workbook_bytes, "test.xlsx")
        from app.modules.leads.services.query import LeadFilters, fetch_filtered

        first_count = len(fetch_filtered(client, ORG, LeadFilters()))

        second_batch = import_service.commit(client, ORG, workbook_bytes, "test.xlsx")
        second_count = len(fetch_filtered(client, ORG, LeadFilters()))

        assert second_count == first_count  # no duplicates
        assert second_batch["rows_inserted"] == 0
        assert second_batch["rows_updated"] == 6

    def test_novo_retorno_column_wins_over_name_prefix(self, workbook_bytes):
        client = _scoped_mock()
        import_service.commit(client, ORG, workbook_bytes, "test.xlsx")
        from app.modules.leads.services.query import LeadFilters, fetch_filtered

        rows = fetch_filtered(client, ORG, LeadFilters())
        bob = next(r for r in rows if r["cliente_nome"] == "Bob")
        assert bob["tipo_lead"] == "retorno"
        alice = next(r for r in rows if r["cliente_nome"] == "Alice")
        assert alice["tipo_lead"] == "novo"

    def test_contact_and_tier_parsed(self, workbook_bytes):
        client = _scoped_mock()
        import_service.commit(client, ORG, workbook_bytes, "test.xlsx")
        from app.modules.leads.services.query import LeadFilters, fetch_filtered

        rows = fetch_filtered(client, ORG, LeadFilters())
        bob = next(r for r in rows if r["cliente_nome"] == "Bob")
        assert bob["contato_tipo"] == "telefone"
        assert bob["contato_norm"] == "11988887777"
        assert bob["anuncio_tier"] == "simples"

    def test_dd_mm_yy_date_parsed(self, workbook_bytes):
        client = _scoped_mock()
        import_service.commit(client, ORG, workbook_bytes, "test.xlsx")
        from app.modules.leads.services.query import LeadFilters, fetch_filtered

        rows = fetch_filtered(client, ORG, LeadFilters())
        daniela = next(r for r in rows if r["cliente_nome"] == "Daniela")
        assert daniela["data_entrada"] == "2024-07-02"

    def test_list_batches_paginated(self, workbook_bytes):
        client = _scoped_mock()
        import_service.commit(client, ORG, workbook_bytes, "test.xlsx")
        rows, total = import_service.list_batches(client, ORG)
        assert total == 1
        assert rows[0]["filename"] == "test.xlsx"
